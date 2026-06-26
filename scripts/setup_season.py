#!/usr/bin/env python3
"""
setup_season.py — script de SETUP, se ejecuta a mano una vez por temporada.

Este script NO toca los ficheros que usa el bot semanal
(data/calendar.json, data/drivers_wrc.json, data/drivers_f1.json).

En su lugar, genera tres ficheros "_nueva_temporada" en la misma carpeta:
    data/calendar_nueva_temporada.json
    data/drivers_wrc_nueva_temporada.json
    data/drivers_f1_nueva_temporada.json

para que los revises y corrijas a mano antes de renombrarlos / sustituir
los originales. Este script SÍ usa red (scraping + API), a diferencia de
check_porra.py que es 100% offline con datos estáticos.

Fuentes:
  - F1: API pública de Jolpica (sucesora de Ergast), devuelve JSON.
  - WRC: scraping ligero de la web oficial — es la parte más frágil,
    por eso el resultado SIEMPRE hay que revisarlo a mano.

Uso:
    python3 scripts/setup_season.py --year 2027
"""

import argparse
import html
import json
import re
import sys
import urllib.request
import urllib.error
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"

FIA_F1_CALENDAR_URL = "https://www.fia.com/events/fia-formula-one-world-championship/season-{year}/formula-one"
FIA_F1_ENTRY_LIST_URL = "https://www.fia.com/events/fia-formula-one-world-championship/season-{year}/{year}-fia-formula-one-world-championship-entry"
FIA_WRC_CALENDAR_URL = "https://www.fia.com/events/world-rally-championship/season-{year}/events-calendar"
WRC_DRIVERS_URL = "https://www.wrc.com/en/teams-and-drivers?rb3TabId=wrc"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,es;q=0.8",
    "Referer": "https://www.wrc.com/",
}


def fetch_url(url: str, timeout: int = 20) -> str:
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        charset = resp.headers.get_content_charset() or "utf-8"
        return resp.read().decode(charset, errors="replace")


def fetch_json(url: str, timeout: int = 20) -> dict:
    raw = fetch_url(url, timeout=timeout)
    return json.loads(raw)


# ──────────────────────────────────────────────────────────────────────
# F1 — scraping de las páginas oficiales de la FIA
# ──────────────────────────────────────────────────────────────────────
#
# Igual que el bloque de WRC más abajo: esto es scraping best-effort.
# La FIA es una fuente más estable que la web comercial de F1/WRC, pero
# sigue siendo HTML que puede cambiar de estructura. El resultado SIEMPRE
# se revisa a mano antes de sustituir nada.

def iso2_to_flag(code: str) -> str:
    """Convierte un código de país ISO 3166-1 alpha-2 (ej. 'AT') en su
    emoji de bandera correspondiente, combinando los dos caracteres
    regionales Unicode."""
    code = code.strip().upper()
    if len(code) != 2 or not code.isalpha():
        return ""
    return "".join(chr(0x1F1E6 + ord(c) - ord("A")) for c in code)


# Mapeo ISO 3166-1 alpha-2 -> nombre de país en español, para los países
# habituales del calendario F1/WRC. Si aparece un código nuevo que no esté
# aquí, se usa el código de 3 letras que da la FIA como mejor esfuerzo,
# y conviene añadirlo a este diccionario en la revisión manual.
COUNTRY_NAMES_ES = {
    "AU": "Australia", "CN": "China", "JP": "Japón", "BH": "Baréin",
    "SA": "Arabia Saudí", "US": "Estados Unidos", "CA": "Canadá",
    "MC": "Mónaco", "ES": "España", "AT": "Austria", "GB": "Reino Unido",
    "BE": "Bélgica", "HU": "Hungría", "NL": "Países Bajos", "IT": "Italia",
    "AZ": "Azerbaiyán", "SG": "Singapur", "MX": "México", "BR": "Brasil",
    "QA": "Catar", "AE": "Emiratos Árabes Unidos",
    "MA": "Marruecos", "SE": "Suecia", "KE": "Kenia", "HR": "Croacia",
    "PT": "Portugal", "GR": "Grecia", "EE": "Estonia", "FI": "Finlandia",
    "PY": "Paraguay", "CL": "Chile", "LV": "Letonia", "LU": "Luxemburgo",
    "NZ": "Nueva Zelanda",
}


def iso2_to_country_name(code: str, fallback: str = "") -> str:
    code = code.strip().upper()
    return COUNTRY_NAMES_ES.get(code, fallback)


# formula1.com/en/drivers da el país en inglés ("Flag of Great Britain"),
# a diferencia de la FIA que da un código ISO2. Mapeo best-effort para
# los países habituales de la parrilla; si aparece uno nuevo que no esté
# aquí, se deja el texto en inglés tal cual para revisión manual.
COUNTRY_NAMES_EN_TO_ES = {
    "Great Britain": "Reino Unido", "United Kingdom": "Reino Unido",
    "Italy": "Italia", "Monaco": "Mónaco", "Netherlands": "Países Bajos",
    "France": "Francia", "Germany": "Alemania", "Spain": "España",
    "Australia": "Australia", "Mexico": "México", "Canada": "Canadá",
    "Finland": "Finlandia", "Thailand": "Tailandia", "Argentina": "Argentina",
    "Brazil": "Brasil", "New Zealand": "Nueva Zelanda",
}


def fetch_fia_f1_calendar(year: int) -> list[dict]:
    url = FIA_F1_CALENDAR_URL.format(year=year)
    print(f"[F1] Descargando calendario desde FIA: {url}")
    try:
        page_html = fetch_url(url)
    except urllib.error.URLError as exc:
        print(f"[F1] AVISO: no se pudo descargar el calendario ({exc}). "
              f"Tendrás que rellenarlo a mano.", file=sys.stderr)
        return []

    # Patrón confirmado contra el HTML real de fia.com (estructura "inner"
    # con date/event/country). Carreras activas van envueltas en <a href>;
    # las canceladas ("Called Off") no llevan ese enlace, así que el grupo
    # opcional de la URL nos sirve también para detectarlas y excluirlas.
    # NOTA: la detección anterior se basaba en la presencia de <a href>
    # alrededor del nombre del evento, asumiendo que las carreras
    # canceladas no llevaban enlace. Eso falló en la práctica: lo más
    # probable es que ese enlace lo añada JavaScript tras la carga
    # inicial (lo que se ve al "Inspeccionar" en el navegador es el DOM
    # ya hidratado, no el HTML crudo que descarga urllib). En su lugar,
    # detectamos directamente el texto "Called Off" que la FIA muestra
    # como texto plano junto a los eventos cancelados/aplazados -- eso
    # sí es contenido estático, presente desde el primer HTML.
    pattern = re.compile(
        r'<div class="day">(\d{1,2})</div>\s*'
        r'<div class="month">(\w+)</div>'
        r'.*?'
        r'<div class="event-name cell">([^<]+)</div>\s*'
        r'<div class="event-location">([^<]*)</div>'
        r'(.*?)'
        r'countrycode-(\w{2})'
        r'.*?'
        r'<div class="country-name">([^<]+)</div>',
        re.DOTALL,
    )

    month_map = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
    }

    calendar = []
    skipped_cancelled = 0
    for m in pattern.finditer(page_html):
        day, month_txt, name, location, between_text, country_iso2, country_name = m.groups()

        if "Called Off" in between_text or "Postponed" in between_text:
            # Cancelado/aplazado -- no se va a disputar, no lo incluimos.
            skipped_cancelled += 1
            continue

        month_num = month_map.get(month_txt[:3].title())
        if not month_num:
            continue  # mes no reconocido, lo dejamos fuera para revisión manual

        race_date = f"{year}-{month_num}-{int(day):02d}"

        calendar.append({
            "round": len(calendar) + 1,
            "name": html.unescape(name.strip()),
            "country": iso2_to_country_name(country_iso2, fallback=html.unescape(country_name.strip())),
            "flag": iso2_to_flag(country_iso2),
            # La FIA solo da el día de carrera (domingo); el fin de semana
            # F1 suele ser viernes-domingo, así que "start" es aproximado
            # y conviene revisarlo si hay sesiones especiales (sprint, etc.)
            "start": race_date,
            "end": race_date,
            "sprint": False,  # la FIA no marca el sprint con claridad aquí; revisar a mano
        })

    if skipped_cancelled:
        print(f"[F1] {skipped_cancelled} evento(s) marcados 'Called Off' excluidos del calendario.")

    if not calendar:
        print("[F1] AVISO: el patrón de scraping no encontró nada en la web de la FIA. "
              "Rellena el calendario F1 a mano este año.", file=sys.stderr)
    else:
        print(f"[F1] {len(calendar)} carreras encontradas (revisar 'start' real del finde y 'sprint').")

    return calendar


def fetch_f1_drivers(year: int) -> list[dict]:
    """
    Scraping de formula1.com/en/drivers (la entry list oficial de la FIA
    es una imagen, no texto, así que no es scrapeable -- ver notas en
    el README). Esta página da nombre, equipo y país por piloto; el
    'code' de 3 letras se sugiere a partir del apellido y SIEMPRE hay
    que revisarlo a mano (colisiones, casos especiales tipo "BOT" vs
    "BOR", etc.)
    """
    url = "https://www.formula1.com/en/drivers"
    print(f"[F1] Descargando pilotos desde {url} ...")
    try:
        page_html = fetch_url(url)
    except urllib.error.URLError as exc:
        print(f"[F1] AVISO: no se pudo descargar la lista de pilotos ({exc}). "
              f"Tendrás que rellenarla a mano.", file=sys.stderr)
        return []

    # Patrón confirmado contra el HTML real de formula1.com/en/drivers:
    # nombre y apellido en <p> con clases "display-l-regular"/"display-l-bold",
    # equipo en <p class="...body-xs-semibold...">, país en <title>Flag of X</title>
    pattern = re.compile(
        r'href="(/en/drivers/[^"]+)"'
        r'.*?'
        r'<p class="[^"]*display-l-regular[^"]*"[^>]*>([^<]+)</p>\s*'
        r'<p class="[^"]*display-l-bold[^"]*"[^>]*>([^<]+)</p>'
        r'.*?'
        r'<p class="[^"]*body-xs-semibold[^"]*"[^>]*>([^<]+)</p>'
        r'.*?'
        r'<title>Flag of ([^<]+)</title>',
        re.DOTALL,
    )

    seen_codes: dict[str, int] = {}
    drivers = []
    for m in pattern.finditer(page_html):
        _slug, first_name, last_name, team, country_en = m.groups()
        full_name = html.unescape(f"{first_name.strip()} {last_name.strip()}")

        suggested_code = last_name.strip()[:3].upper()
        seen_codes[suggested_code] = seen_codes.get(suggested_code, 0) + 1

        drivers.append({
            "code": suggested_code,  # SUGERENCIA, revisar colisiones a mano
            "name": full_name,
            "team": team.strip(),
            "country": COUNTRY_NAMES_EN_TO_ES.get(country_en.strip(), country_en.strip()),
        })

    collisions = [code for code, count in seen_codes.items() if count > 1]
    if collisions:
        print(f"[F1] AVISO: códigos sugeridos duplicados, revisar a mano: {collisions}",
              file=sys.stderr)

    if not drivers:
        print("[F1] AVISO: el patrón de scraping no encontró pilotos en formula1.com. "
              "Rellena la lista F1 a mano este año.", file=sys.stderr)
    else:
        print(f"[F1] {len(drivers)} pilotos encontrados (revisar 'code', "
              f"{len(collisions)} colisión(es) detectada(s)).")

    return drivers


# ──────────────────────────────────────────────────────────────────────
# WRC — scraping ligero (FRÁGIL a propósito: solo da un punto de partida)
# ──────────────────────────────────────────────────────────────────────

def fetch_wrc_calendar(year: int) -> list[dict]:
    """
    Scraping del calendario WRC desde la página de la FIA, usando el
    mismo patrón de estructura "inner" confirmado para F1, con la
    diferencia de que WRC publica un rango from-date/to-date en vez
    de un único día de carrera.

    A diferencia de F1, aquí no hay <a href> que distinga eventos
    cancelados de activos, así que cualquier rally "Called Off" que
    pudiera aparecer no se filtra automáticamente: revisar a mano.
    """
    url = FIA_WRC_CALENDAR_URL.format(year=year)
    print(f"[WRC] Descargando calendario desde FIA: {url}")
    try:
        page_html = fetch_url(url)
    except urllib.error.URLError as exc:
        print(f"[WRC] AVISO: no se pudo descargar el calendario ({exc}). "
              f"Tendrás que rellenarlo a mano.", file=sys.stderr)
        return []

    month_map = {
        "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
        "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
        "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
    }

    pattern = re.compile(
        r'<div class="from-date cell">\s*'
        r'<div class="day">(\d{1,2})</div>\s*'
        r'<div class="month">(\w+)</div>\s*'
        r'</div>\s*'
        r'(?:<div class="to-date cell">\s*'
        r'<div class="day">(\d{1,2})</div>\s*'
        r'<div class="month">(\w+)</div>\s*'
        r'</div>)?'
        r'.*?'
        r'<div class="event-name cell">([^<]+)</div>'
        r'.*?'
        r'countrycode-(\w{2})'
        r'.*?'
        r'<div class="country-name">([^<]+)</div>',
        re.DOTALL,
    )

    calendar = []
    for m in pattern.finditer(page_html):
        from_day, from_month_txt, to_day, to_month_txt, name, country_iso2, country_name = m.groups()

        from_month = month_map.get(from_month_txt[:3].title())
        if not from_month:
            continue  # mes no reconocido, fuera para revisión manual

        start = f"{year}-{from_month}-{int(from_day):02d}"

        if to_day and to_month_txt:
            to_month = month_map.get(to_month_txt[:3].title())
            # si el rally cruza fin de año (ej. 30 Dic - 2 Ene), el año
            # del "to" sería el siguiente -- caso raro en WRC, pero por si acaso
            to_year = year + 1 if to_month and int(to_month) < int(from_month) else year
            end = f"{to_year}-{to_month}-{int(to_day):02d}" if to_month else start
        else:
            end = start  # sin rango, un único día

        calendar.append({
            "round": len(calendar) + 1,
            "name": html.unescape(name.strip()),
            "country": iso2_to_country_name(country_iso2, fallback=html.unescape(country_name.strip())),
            "flag": iso2_to_flag(country_iso2),
            "start": start,
            "end": end,
        })

    if not calendar:
        print("[WRC] AVISO: el patrón de scraping no encontró nada en la web de la FIA. "
              "Es probable que haya cambiado de estructura, o que no haya eventos "
              "'from-date'/'to-date' en este formato. Rellena el calendario WRC a mano.",
              file=sys.stderr)
    else:
        print(f"[WRC] {len(calendar)} rallies encontrados "
              "(revisar si hay alguno cancelado/aplazado: no se detecta automáticamente en WRC).")

    return calendar


def fetch_wrc_drivers() -> list[dict]:
    """
    Scraping de la tabla de pilotos Rally1 en wrc.com/en/teams-and-drivers.

    AVISO HONESTO: esta web usa Web Components personalizados (Cosmos,
    el sistema de diseño de Red Bull) que se "hidratan" con JavaScript.
    El HTML que descarga urllib (sin ejecutar JS) puede llegar
    prácticamente vacío, en cuyo caso esta función no encontrará nada
    y caerá al aviso de "rellenar a mano" -- que es el comportamiento
    esperado y aceptado para este caso, no un fallo del patrón en sí.
    """
    print(f"[WRC] Descargando pilotos desde {WRC_DRIVERS_URL} ...")
    try:
        page_html = fetch_url(WRC_DRIVERS_URL)
    except urllib.error.URLError as exc:
        print(f"[WRC] AVISO: no se pudo descargar pilotos ({exc}). "
              f"Tendrás que rellenarlo a mano.", file=sys.stderr)
        return []

    # Patrón sobre la estructura confirmada por inspección manual del
    # navegador. IMPORTANTE: la página lista pilotos Y copilotos juntos
    # en tarjetas "driver-couple-card". Anclamos el patrón al wrapper
    # class="driver-couple-card__driver" (no "...__co-driver") para
    # quedarnos solo con los pilotos.
    pattern = re.compile(
        r'class="driver-couple-card__driver">'
        r'.*?'
        r'class="driver-view__name"[^>]*>([^<]+)</cosmos-title-[\d.-]+>'
        r'.*?'
        r'<img[^>]*alt="([^"]+)"[^>]*class="driver-view__flag"',
        re.DOTALL,
    )
    matches = pattern.findall(page_html)

    drivers = []
    for name, country in matches:
        name = html.unescape(name.strip())
        surname = name.split()[-1]
        suggested_code = surname[:3].upper()  # sugerencia, revisar colisiones a mano
        drivers.append({
            "code": suggested_code,
            "name": name,
            "team": "",  # esta página no muestra el equipo junto al piloto; rellenar a mano
            "country": html.unescape(country.strip()),
        })

    if not drivers:
        print("[WRC] AVISO: el patrón de scraping no encontró pilotos. Es muy probable que "
              "wrc.com cargue esta lista por JavaScript (Web Components) y urllib no pueda "
              "verla sin ejecutar JS. Rellena la lista WRC a mano este año.", file=sys.stderr)
    else:
        print(f"[WRC] {len(drivers)} pilotos encontrados (revisar 'code' y 'team', "
              "este último siempre vacío en esta fuente).")

    return drivers


# ──────────────────────────────────────────────────────────────────────
# Escritura de ficheros de revisión
# ──────────────────────────────────────────────────────────────────────

def write_review_json(filename: str, payload) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    path = DATA_DIR / filename
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"  -> escrito {path}")
    return path


# ──────────────────────────────────────────────────────────────────────
# Excel/Numbers — calendario combinado, ordenado por fecha, con color
# ──────────────────────────────────────────────────────────────────────

def write_calendar_xlsx(wrc_calendar: list[dict], f1_calendar: list[dict], year: int) -> Path:
    """
    Genera un .xlsx (importable en Excel o Numbers) con columnas
    "FECHA" (fecha de la carrera, formato "MMM - dd") y "PRUEBA"
    (bandera + nombre), ordenado por fecha de inicio ascendente.
    Filas WRC en azul/blanco/negrita, filas F1 en rojo/blanco/negrita.
    """
    from datetime import date as date_cls
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = []
    for race in wrc_calendar:
        rows.append({
            "categoria": "WRC",
            "fecha": date_cls.fromisoformat(race["start"]),
            "prueba": f"{race.get('flag', '')} {race['name']}".strip(),
        })
    for race in f1_calendar:
        rows.append({
            "categoria": "F1",
            "fecha": date_cls.fromisoformat(race["start"]),
            "prueba": f"{race.get('flag', '')} {race['name']}".strip(),
        })

    rows.sort(key=lambda r: r["fecha"])

    wb = Workbook()
    sheet = wb.active
    sheet.title = f"Calendario {year}"

    sheet["A1"] = "FECHA"
    sheet["B1"] = "PRUEBA"
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 45

    wrc_fill = PatternFill("solid", start_color="1F4E96", end_color="1F4E96")  # azul
    f1_fill = PatternFill("solid", start_color="C00000", end_color="C00000")   # rojo
    white_bold = Font(bold=True, color="FFFFFF")

    for i, row in enumerate(rows, start=2):
        fill = wrc_fill if row["categoria"] == "WRC" else f1_fill

        fecha_cell = sheet.cell(row=i, column=1, value=row["fecha"])
        fecha_cell.font = white_bold
        fecha_cell.fill = fill
        fecha_cell.alignment = Alignment(vertical="center")
        fecha_cell.number_format = "MMM - dd"

        prueba_cell = sheet.cell(row=i, column=2, value=row["prueba"])
        prueba_cell.font = white_bold
        prueba_cell.fill = fill
        prueba_cell.alignment = Alignment(vertical="center")

    path = DATA_DIR / f"calendario_{year}.xlsx"
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  -> escrito {path} ({len(rows)} pruebas, ordenadas por fecha)")
    return path


def main():
    parser = argparse.ArgumentParser(
        description="Genera ficheros _nueva_temporada para revisión manual."
    )
    parser.add_argument("--year", type=int, required=True,
                         help="Año de la nueva temporada, p.ej. 2027")
    parser.add_argument("--skip-wrc", action="store_true",
                         help="No intentar el scraping de WRC (solo F1)")
    parser.add_argument("--skip-f1", action="store_true",
                         help="No consultar la API de F1 (solo WRC)")
    args = parser.parse_args()

    print(f"=== Setup temporada {args.year} ===\n")

    if not args.skip_f1:
        try:
            f1_calendar = fetch_fia_f1_calendar(args.year)
            f1_drivers = fetch_f1_drivers(args.year)
        except (urllib.error.URLError, json.JSONDecodeError) as exc:
            print(f"[F1] ERROR obteniendo datos: {exc}", file=sys.stderr)
            f1_calendar, f1_drivers = [], []
    else:
        print("[F1] Saltado por --skip-f1")
        f1_calendar, f1_drivers = [], []

    print()

    if not args.skip_wrc:
        wrc_calendar = fetch_wrc_calendar(args.year)
        wrc_drivers = fetch_wrc_drivers()
    else:
        print("[WRC] Saltado por --skip-wrc")
        wrc_calendar, wrc_drivers = [], []

    print("\n=== Escribiendo ficheros de revisión ===")

    calendar_payload = {
        "wrc": wrc_calendar,
        "f1": f1_calendar,
        "_notes": (
            f"Generado por setup_season.py para {args.year} a partir de las páginas "
            "de la FIA (fia.com). REVISAR A MANO antes de sustituir data/calendar.json: "
            "comprobar fechas, banderas, países, y cualquier carrera "
            "cancelada/aplazada (sponsors, conflictos, etc. no se detectan "
            "automáticamente)."
        ),
    }
    write_review_json("calendar_nueva_temporada.json", calendar_payload)
    write_review_json("drivers_wrc_nueva_temporada.json", wrc_drivers)
    write_review_json("drivers_f1_nueva_temporada.json", f1_drivers)

    write_calendar_xlsx(wrc_calendar, f1_calendar, args.year)

    print(
        "\nListo. Revisa y corrige a mano los 3 ficheros en data/*_nueva_temporada.json.\n"
        "Cuando estén correctos, renómbralos (quitando '_nueva_temporada') para "
        "sustituir a los que usa el bot semanal:\n"
        "    data/calendar_nueva_temporada.json   -> data/calendar.json\n"
        "    data/drivers_wrc_nueva_temporada.json -> data/drivers_wrc.json\n"
        "    data/drivers_f1_nueva_temporada.json  -> data/drivers_f1.json\n"
    )


if __name__ == "__main__":
    main()
